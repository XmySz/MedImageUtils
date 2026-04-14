from pathlib import Path
from openpyxl import load_workbook, Workbook

excel_path = Path(r"D:\Data\Jmszxyy\骨松四分类\Dataset\茂名市人民医院_补充\茂名市人民医院_补充.xlsx")
image_root = Path(r"D:\Data\Jmszxyy\骨松四分类\Dataset\茂名市人民医院_补充\Original")
output_path = Path(r"D:\Data\Jmszxyy\骨松四分类\Dataset\茂名市人民医院_补充\missing_hospital_ids.xlsx")

wb = load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]

headers = {}
for col in range(1, ws.max_column + 1):
    value = ws.cell(1, col).value
    if value is not None:
        headers[str(value).strip()] = col

if "住院号" not in headers:
    raise ValueError("表格中未找到“住院号”列，请检查表头。")

id_col = headers["住院号"]
name_col = headers.get("姓名", None)

rows_data = []
for row in range(2, ws.max_row + 1):
    v = ws.cell(row, id_col).value
    if v is None:
        continue
    hospital_id = str(v).strip()
    if not hospital_id:
        continue

    name = ""
    if name_col is not None:
        nv = ws.cell(row, name_col).value
        name = "" if nv is None else str(nv).strip()

    rows_data.append((hospital_id, name))

# 去重
seen = set()
unique_rows = []
for hospital_id, name in rows_data:
    if hospital_id not in seen:
        seen.add(hospital_id)
        unique_rows.append((hospital_id, name))

existing_ids = set()
for file_path in image_root.rglob("*"):
    if not file_path.is_file():
        continue

    stem = file_path.stem.strip()
    possible_id = stem.split("_")[0].strip() if "_" in stem else stem
    if possible_id:
        existing_ids.add(possible_id)

missing_rows = [(hid, name) for hid, name in unique_rows if hid not in existing_ids]

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "缺失胸片"

out_ws.append(["住院号", "姓名"])
for hid, name in missing_rows:
    out_ws.append([hid, name])

out_wb.save(output_path)

print(f"表格住院号总数: {len(unique_rows)}")
print(f"缺失胸片的住院号数: {len(missing_rows)}")
print(f"结果已保存到: {output_path}")
